import os
import math
from datetime import datetime, timezone
import logging
logging.basicConfig(level=logging.DEBUG)

import io
import csv as csv_module
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import text
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from db import Base, engine, get_db
from models import Tournament, Boat, RuleConfig, Race, RaceResult, Series, RankingProfile, User, TournamentMember
from schemas import (
    TournamentCreate,
    TournamentUpdate,
    TournamentOut,
    BoatCreate,
    BoatOut,
    BoatBulkItem,
    BoatBulkUpdate,
    RuleConfigUpdate,
    RuleConfigOut,
    RaceCreate,
    RaceUpdate,
    RaceOut,
    RaceResultInput,
    RaceResultOut,
    StandingRow,
    SeriesOut,
    RankingProfileOut,
    StandingSection,
    StandingsResponse,
    StandingsV3Response,
    UserOut,
    InviteRequest,
    InviteResponse,
    TournamentMemberOut,
    AddMemberRequest,
)
from auth import get_current_user, require_admin, check_tournament_access, get_supabase, AUTH_ENABLED

try:
    Base.metadata.create_all(bind=engine)
    print("[main] create_all OK", flush=True)
except Exception as e:
    print(f"[main] create_all FAILED: {type(e).__name__}: {e}", flush=True)

# カラム追加マイグレーション — 1文ずつ独立した try/except で囲み、
# 1つが失敗しても他のマイグレーションに影響しないようにする。
_MIGRATIONS = [
    "ALTER TABLE tournaments ADD COLUMN IF NOT EXISTS owner_id TEXT",
    "ALTER TABLE tournament_members ADD COLUMN IF NOT EXISTS role TEXT NOT NULL DEFAULT 'editor'",
    "ALTER TABLE boats ADD COLUMN IF NOT EXISTS entry_number INTEGER",
    "ALTER TABLE boats ADD COLUMN IF NOT EXISTS helmsman_name2 TEXT",
    "ALTER TABLE boats ADD COLUMN IF NOT EXISTS helmsman_name3 TEXT",
    "ALTER TABLE boats ADD COLUMN IF NOT EXISTS crew_name2 TEXT",
    "ALTER TABLE boats ADD COLUMN IF NOT EXISTS crew_name3 TEXT",
    # PostgreSQL専用: NOT NULL 制約を解除（SQLiteでは非対応だが本番はPostgreSQL）
    "ALTER TABLE boats ALTER COLUMN boat_number DROP NOT NULL",
    "ALTER TABLE boats ALTER COLUMN organization_name DROP NOT NULL",
    "ALTER TABLE races ADD COLUMN IF NOT EXISTS race_date TEXT",
    "ALTER TABLE races ADD COLUMN IF NOT EXISTS weather TEXT",
    "ALTER TABLE races ADD COLUMN IF NOT EXISTS wind_direction TEXT",
    "ALTER TABLE races ADD COLUMN IF NOT EXISTS wind_speed TEXT",
    "ALTER TABLE races ADD COLUMN IF NOT EXISTS start_time TEXT",
    "ALTER TABLE races ADD COLUMN IF NOT EXISTS finish_time_top TEXT",
    "ALTER TABLE races ADD COLUMN IF NOT EXISTS finish_time_last TEXT",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS nsc_rule TEXT NOT NULL DEFAULT 'STARTERS_PLUS_1'",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS dne_rule TEXT NOT NULL DEFAULT 'STARTERS_PLUS_1'",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS custom_result_codes TEXT",
    "ALTER TABLE tournaments ADD COLUMN IF NOT EXISTS deleted_at TEXT",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS team_cut_method TEXT NOT NULL DEFAULT 'individual'",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS overall_tie_method TEXT NOT NULL DEFAULT 'kanto'",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS tie_fallback_extended BOOLEAN NOT NULL DEFAULT true",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS tie_use_excluded_scores BOOLEAN NOT NULL DEFAULT true",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS dne_score_method TEXT NOT NULL DEFAULT 'plus_one'",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS sp_method TEXT NOT NULL DEFAULT 'dsq'",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS use_appendix_t BOOLEAN NOT NULL DEFAULT true",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS same_school_rule BOOLEAN NOT NULL DEFAULT false",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS min_races_to_complete INTEGER NOT NULL DEFAULT 1",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS fleet_split BOOLEAN NOT NULL DEFAULT false",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS fleet_split_method TEXT NOT NULL DEFAULT 'own'",
    "ALTER TABLE rule_configs ADD COLUMN IF NOT EXISTS preset_template TEXT NOT NULL DEFAULT 'custom'",
]

for _sql in _MIGRATIONS:
    try:
        with engine.connect() as _conn:
            _conn.execute(text(_sql))
            _conn.commit()
    except Exception as _e:
        print(f"[main] migration skip ({_e.__class__.__name__}): {_sql[:60]}", flush=True)

app = FastAPI()

# ALLOWED_ORIGINS 環境変数があればそれを使い、なければワイルドカードを許可
_raw_origins = os.environ.get("ALLOWED_ORIGINS", "")
allow_origins_list = [o.strip() for o in _raw_origins.split(",") if o.strip()] or ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins_list,
    allow_credentials=allow_origins_list != ["*"],
    allow_methods=["*"],
    allow_headers=["Content-Type", "Authorization"],
)

def get_entries_count(tournament_id: int, db: Session, boat_class: str | None = None) -> int:
    q = db.query(Boat).filter(Boat.tournament_id == tournament_id)
    if boat_class is not None:
        q = q.filter(Boat.boat_class == boat_class)
    return q.count()


def get_starters_count(payload: list[RaceResultInput], boat_class_map: dict[int, str] | None = None, boat_class: str | None = None) -> int:
    non_starters = {"DNC", "DNS"}
    if boat_class is None or boat_class_map is None:
        return sum(1 for item in payload if item.result_code not in non_starters)
    return sum(
        1 for item in payload
        if item.result_code not in non_starters
        and boat_class_map.get(item.boat_id, "") == boat_class
    )


def apply_scoring_rule(rule_name: str, entries_count: int, starters_count: int) -> int:
    if rule_name == "STARTERS_PLUS_1":
        return starters_count + 1
    if rule_name == "ENTRIES_PLUS_1":
        return entries_count + 1

    raise HTTPException(status_code=400, detail=f"Unknown scoring rule: {rule_name}")


def parse_class_config(class_config: str | None) -> list[str]:
    """
    "470,SNIPE"     -> ["470", "SNIPE"]
    "OTHER:ILCA"    -> ["OTHER:ILCA"]
    "470,OTHER:ILCA"-> ["470", "OTHER:ILCA"]
    None / ""       -> []
    """
    if not class_config:
        return []
    return [c.strip() for c in class_config.split(",") if c.strip()]


def class_display_name(entry: str) -> str:
    """class_config エントリをフィルタ用クラス名に変換する。
    "OTHER:ILCA" -> "ILCA", "470" -> "470"
    """
    if entry.startswith("OTHER:"):
        return entry[len("OTHER:"):]
    return entry


def calculate_points_for_result(
    item: RaceResultInput,
    rule_config: RuleConfig,
    entries_count: int,
    starters_count: int,
) -> int:
    code = item.result_code
    fp   = item.finish_position

    # 通常フィニッシュ
    if code == "OK":
        if fp is None:
            raise HTTPException(status_code=400, detail="OK result requires finish_position")
        return fp

    # 手動得点コード（RDG: 裁定による救済、DPI: 任意裁量ペナルティ）
    if code in ("RDG", "DPI"):
        return item.manual_points if item.manual_points is not None else 0

    # 標準ペナルティテーブルコード
    rule_map = {
        "DNC": rule_config.dnc_rule,   # 出場エリアに来なかった
        "DNS": rule_config.dns_rule,   # スタートしなかった
        "OCS": rule_config.ocs_rule,   # スタートライン越え
        "DNF": rule_config.dnf_rule,   # フィニッシュしなかった
        "RET": rule_config.ret_rule,   # リタイア
        "DSQ": rule_config.dsq_rule,   # 失格
        "UFD": rule_config.ufd_rule,   # Uフラッグ失格
        "BFD": rule_config.bfd_rule,   # 黒旗失格
        "NSC": rule_config.dsq_rule,   # コースを航走しなかった（DSQ同等）
    }
    if code in rule_map:
        return apply_scoring_rule(rule_map[code], entries_count, starters_count)

    # DNE: 除外不能な失格 — dne_score_method に応じてエントリー艇数 +1 または +5
    if code == "DNE":
        if getattr(rule_config, "dne_score_method", "plus_one") == "plus_five":
            return entries_count + 5
        return entries_count + 1

    # 着順ベースのペナルティコード（fp 必須）
    if fp is None:
        raise HTTPException(status_code=400, detail=f"{code} requires finish_position")

    if code == "STP":
        sp_method = getattr(rule_config, "sp_method", "dsq")
        if sp_method == "add_one":
            return fp + 1
        return apply_scoring_rule(rule_config.dsq_rule, entries_count, starters_count)

    if code in ("SCP", "ARB"):               # 各種ペナルティ ×1.3 切り上げ
        return math.ceil(fp * 1.3)

    if code == "PRP":                        # 付則T — 無効時はエラー
        if not getattr(rule_config, "use_appendix_t", True):
            raise HTTPException(status_code=400, detail="PRP コードは無効です（付則T が適用されていません）")
        return math.ceil(fp * 1.3)

    if code == "ZFP":                        # 規則30.2 ×1.2 切り上げ
        return math.ceil(fp * 1.2)

    raise HTTPException(status_code=400, detail=f"Unsupported result_code: {code}")

def calculate_individual_standings(tournament_id: int, db: Session):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    boats = (
        db.query(Boat)
        .filter(Boat.tournament_id == tournament_id)
        .order_by(Boat.id.asc())
        .all()
    )

    races = (
        db.query(Race)
        .filter(Race.tournament_id == tournament_id)
        .order_by(Race.race_number.asc())
        .all()
    )

    rule_config = db.query(RuleConfig).filter(RuleConfig.tournament_id == tournament_id).first()
    if rule_config is None:
        raise HTTPException(status_code=404, detail="RuleConfig not found")

    results = (
        db.query(RaceResult)
        .join(Race, Race.id == RaceResult.race_id)
        .filter(Race.tournament_id == tournament_id)
        .all()
    )

    result_map = {}
    for result in results:
        result_map[(result.race_id, result.boat_id)] = result

    completed_races = len(races)

    standings = []
    for boat in boats:
        race_points = []
        valid_points = []

        for race in races:
            result = result_map.get((race.id, boat.id))
            if result is None or result.points is None:
                race_points.append(None)
            else:
                race_points.append(result.points)
                valid_points.append(result.points)

        total_points = sum(valid_points)

        discarded_points = []
        net_points = total_points

        if (
            rule_config.discard_enabled == 1
            and rule_config.discard_start_race_count is not None
            and rule_config.discard_count is not None
            and completed_races >= rule_config.discard_start_race_count
            and len(valid_points) > 0
        ):
            sorted_desc = sorted(valid_points, reverse=True)
            discarded_points = sorted_desc[: rule_config.discard_count]
            net_points = total_points - sum(discarded_points)

        standings.append(
            {
                "boat_id": boat.id,
                "boat_number": boat.boat_number,
                "sail_number": boat.sail_number,
                "organization_name": boat.organization_name,
                "race_points": race_points,
                "total_points": total_points,
                "discarded_points": discarded_points,
                "net_points": net_points,
                "rank": 0,
            }
        )

    standings.sort(key=lambda x: (
        0 if any(p is not None for p in x["race_points"]) else 1,
        x["net_points"],
        x["total_points"],
        x["boat_number"] or ""
    ))

    for i, row in enumerate(standings, start=1):
        row["rank"] = i

    return standings


def calculate_team_standings(tournament_id: int, db: Session, team_size: int):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    boats = (
        db.query(Boat)
        .filter(Boat.tournament_id == tournament_id)
        .order_by(Boat.id.asc())
        .all()
    )

    races = (
        db.query(Race)
        .filter(Race.tournament_id == tournament_id)
        .order_by(Race.race_number.asc())
        .all()
    )

    results = (
        db.query(RaceResult)
        .join(Race, Race.id == RaceResult.race_id)
        .filter(Race.tournament_id == tournament_id)
        .all()
    )

    boat_map = {boat.id: boat for boat in boats}

    # race_id ごとに team ごとの得点一覧をためる
    race_team_points = {}
    for race in races:
        race_team_points[race.id] = {}

    for result in results:
        if result.points is None:
            continue

        boat = boat_map.get(result.boat_id)
        if boat is None:
            continue

        team_name = boat.team_name if boat.team_name else boat.organization_name
        if not team_name:
            continue

        if team_name not in race_team_points[result.race_id]:
            race_team_points[result.race_id][team_name] = []

        race_team_points[result.race_id][team_name].append(result.points)

    standings = []
    all_team_names = set()

    for boat in boats:
        team_name = boat.team_name if boat.team_name else boat.organization_name
        if team_name:
            all_team_names.add(team_name)

    for team_name in sorted(all_team_names):
        race_points = []
        total_points = 0

        for race in races:
            points_list = sorted(race_team_points[race.id].get(team_name, []))
            if len(points_list) == 0:
                race_points.append(None)
            else:
                adopted = sum(points_list[:team_size])
                race_points.append(adopted)
                total_points += adopted

        is_incomplete = any(
            0 < len(race_team_points[race.id].get(team_name, [])) < team_size
            for race in races
        )
        has_any = any(p is not None for p in race_points)
        standings.append(
            {
                "boat_id": 0,
                "boat_number": "",
                "sail_number": "",
                "organization_name": team_name,
                "race_points": race_points,
                "total_points": total_points,
                "discarded_points": [],
                "net_points": total_points,
                "rank": 0,
                "_sort_group": 0 if (has_any and not is_incomplete) else (1 if has_any else 2),
            }
        )

    standings.sort(key=lambda x: (x["_sort_group"], x["net_points"], x["organization_name"]))
    for row in standings:
        del row["_sort_group"]

    for i, row in enumerate(standings, start=1):
        row["rank"] = i

    return standings


def calculate_multi_group_hybrid_standings(tournament_id: int, db: Session):
    raise HTTPException(status_code=501, detail="MULTI_GROUP_HYBRID standings not implemented yet")


def calculate_standings(tournament_id: int, db: Session):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    if tournament.event_template == "INDIVIDUAL":
        return calculate_individual_standings(tournament_id, db)

    if tournament.event_template == "TEAM_3_BOATS":
        return calculate_team_standings(tournament_id, db, team_size=3)

    if tournament.event_template == "TEAM_4_BOATS":
        return calculate_team_standings(tournament_id, db, team_size=4)

    if tournament.event_template == "MULTI_GROUP_HYBRID":
        return calculate_multi_group_hybrid_standings(tournament_id, db)

    raise HTTPException(status_code=400, detail=f"Unsupported event_template: {tournament.event_template}")

def get_race_result_details_by_boat(tournament_id: int, db: Session):
    races = (
        db.query(Race)
        .filter(Race.tournament_id == tournament_id)
        .order_by(Race.race_number.asc())
        .all()
    )

    results = (
        db.query(RaceResult)
        .join(Race, Race.id == RaceResult.race_id)
        .filter(Race.tournament_id == tournament_id)
        .all()
    )

    result_map = {}
    for result in results:
        result_map[(result.race_id, result.boat_id)] = {
            "finish_position": result.finish_position,
            "result_code": result.result_code,
            "points": result.points,
        }

    return races, result_map

def _parse_class_config(cfg: str | None) -> list[str]:
    if not cfg:
        return []
    result = []
    for part in cfg.split(","):
        part = part.strip()
        if part.startswith("OTHER:"):
            result.append(part[6:])
        elif part:
            result.append(part)
    return result


def build_standings_workbook(tournament_id: int, db: Session) -> Workbook:
    from openpyxl.utils import get_column_letter

    now = datetime.now()

    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    is_team   = tournament.event_template in ("TEAM_3_BOATS", "TEAM_4_BOATS", "MULTI_GROUP_HYBRID")
    team_size = 4 if tournament.event_template == "TEAM_4_BOATS" else 3

    all_boats = (
        db.query(Boat)
        .filter(Boat.tournament_id == tournament_id)
        .order_by(Boat.id.asc())
        .all()
    )
    races, race_result_map = get_race_result_details_by_boat(tournament_id, db)

    classes = _parse_class_config(tournament.class_config)

    # ─── Shared styles ────────────────────────────────────────────────────────
    navy_fill  = PatternFill("solid", fgColor="1F4E78")
    gray_fill  = PatternFill("solid", fgColor="F0F4F8")
    hdr_font   = Font(color="FFFFFF", bold=True, size=9)
    bold9      = Font(bold=True, size=9)
    norm9      = Font(size=9)
    center_al  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al    = Alignment(horizontal="left",   vertical="center")
    right_al   = Alignment(horizontal="right",  vertical="center")
    thin_side  = Side(style="thin",   color="AAAAAA")
    med_side   = Side(style="medium", color="000000")

    def tb():   # thin border
        return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def race_box_border(pos: str):
        """pos: 'left' | 'mid' | 'right'  — thick left/right edges around a race group"""
        l = med_side  if pos == "left"  else thin_side
        r = med_side  if pos == "right" else thin_side
        return Border(left=l, right=r, top=thin_side, bottom=thin_side)

    def race_box_top(pos: str):
        l = med_side  if pos == "left"  else thin_side
        r = med_side  if pos == "right" else thin_side
        return Border(left=l, right=r, top=med_side, bottom=thin_side)

    def ap(cell, fill=None, font=None, alignment=None, border=None):
        if fill:      cell.fill      = fill
        if font:      cell.font      = font
        if alignment: cell.alignment = alignment
        if border:    cell.border    = border

    def apply_outer_border(ws, r1, r2, c1, c2):
        from openpyxl.cell.cell import MergedCell
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue  # 結合セル内部はスキップ（top-left で管理）
                b = cell.border
                cell.border = Border(
                    top=med_side    if row == r1 else b.top,
                    bottom=med_side if row == r2 else b.bottom,
                    left=med_side   if col == c1 else b.left,
                    right=med_side  if col == c2 else b.right,
                )

    # n_fixed: 順位 Entry 大学 艇体 セール Helm1 Helm2 Helm3 Crew1 Crew2 Crew3
    N_FIXED = 11

    date_str = ""
    if tournament.start_date:
        date_str = tournament.start_date
        if tournament.end_date and tournament.end_date != tournament.start_date:
            date_str += f"〜{tournament.end_date}"

    # ─── Sheet writer ──────────────────────────────────────────────────────────
    def write_sheet(wb: Workbook, sheet_title: str, boats: list, sheet_class: str | None):
        ws = wb.create_sheet(title=sheet_title)
        n_races  = len(races)
        # tail: 艇合計 + 大学合計 for team, 合計得点 for individual
        n_tail   = 2 if is_team else 1
        total_cols = N_FIXED + n_races * 3 + n_tail

        # ── Row 1: title ──────────────────────────────────────────────────────
        class_label = sheet_class or tournament.class_name or ""
        title_text  = tournament.name
        if class_label:
            title_text += f"  {class_label}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        c = ws.cell(row=1, column=1, value=title_text)
        c.font = Font(bold=True, size=13); c.alignment = left_al

        # ── Row 2: timestamp (top-right, 8pt gray) ────────────────────────────
        ts_cell = ws.cell(row=2, column=total_cols,
                          value=f"更新日時: {now.strftime('%Y年%m月%d日 %H:%M')}")
        ts_cell.font = Font(size=8, color="888888")
        ts_cell.alignment = Alignment(horizontal="right", vertical="center")

        # ── Rows 3-4: double-row header ───────────────────────────────────────
        H1, H2 = 3, 4  # header rows

        fixed_hdrs = ["順位", "ｴﾝﾄﾘｰ\nNo", "大学名", "艇体\nNo", "セール\nNo",
                      "Helm 1", "Helm 2", "Helm 3", "Crew 1", "Crew 2", "Crew 3"]
        for ci, h in enumerate(fixed_hdrs, 1):
            # merge H1 and H2 for fixed columns
            ws.merge_cells(start_row=H1, start_column=ci, end_row=H2, end_column=ci)
            c = ws.cell(row=H1, column=ci, value=h)
            ap(c, fill=navy_fill, font=hdr_font, alignment=center_al, border=tb())
            ws.cell(row=H2, column=ci).border = tb()

        # Race group headers
        col = N_FIXED + 1
        for ri, race in enumerate(races):
            rn = race.race_number
            c1, c2, c3 = col, col + 1, col + 2
            # H1: "第NR" merged across 3 cols
            ws.merge_cells(start_row=H1, start_column=c1, end_row=H1, end_column=c3)
            c = ws.cell(row=H1, column=c1, value=f"第{rn}Ｒ")
            ap(c, fill=navy_fill, font=hdr_font, alignment=center_al,
               border=Border(left=med_side, right=med_side, top=med_side, bottom=thin_side))
            # H2: 着順 / 順位 / 得点
            for sub_col, sub_hdr in zip([c1, c2, c3], ["着順", "順位", "得点"]):
                pos = "left" if sub_col == c1 else ("right" if sub_col == c3 else "mid")
                c = ws.cell(row=H2, column=sub_col, value=sub_hdr)
                ap(c, fill=navy_fill, font=hdr_font, alignment=center_al, border=race_box_border(pos))
            col += 3

        # Tail header(s)
        if is_team:
            for tail_h in ("艇合計", "大学合計"):
                ws.merge_cells(start_row=H1, start_column=col, end_row=H2, end_column=col)
                c = ws.cell(row=H1, column=col, value=tail_h)
                ap(c, fill=navy_fill, font=hdr_font, alignment=center_al, border=tb())
                ws.cell(row=H2, column=col).border = tb()
                col += 1
        else:
            ws.merge_cells(start_row=H1, start_column=col, end_row=H2, end_column=col)
            c = ws.cell(row=H1, column=col, value="合計得点")
            ap(c, fill=navy_fill, font=hdr_font, alignment=center_al, border=tb())
            ws.cell(row=H2, column=col).border = tb()

        ws.row_dimensions[H1].height = 18
        ws.row_dimensions[H2].height = 18

        # ── Data rows ─────────────────────────────────────────────────────────
        data_start   = H2 + 1
        current_row  = data_start

        def write_boat_row(row_num: int, rank_val, univ_val, boat, boat_total, team_total_val,
                           is_first_in_team: bool, boat_idx_in_team: int = 0, team_boat_count: int = 1):
            top_bdr = med_side if is_first_in_team else thin_side

            # 大学名: チーム内の中段の行にのみ表示（中央揃え）
            mid_idx = (team_boat_count - 1) // 2
            show_univ = univ_val if boat_idx_in_team == mid_idx else ""

            def wc(col_i, val, al=None, fnt=None):
                c = ws.cell(row=row_num, column=col_i, value=val)
                ap(c, font=fnt or norm9, alignment=al or left_al,
                   border=Border(left=thin_side, right=thin_side,
                                 top=top_bdr, bottom=thin_side))

            wc(1,  rank_val if rank_val != "" else "", center_al, bold9 if rank_val != "" else norm9)
            wc(2,  boat.entry_number if boat.entry_number is not None else "", center_al)
            wc(3,  show_univ, center_al, bold9 if show_univ else norm9)
            wc(4,  boat.boat_number or "", center_al)
            wc(5,  boat.sail_number or "", center_al)
            wc(6,  boat.helmsman_name or "", left_al)
            wc(7,  boat.helmsman_name2 or "", left_al)
            wc(8,  boat.helmsman_name3 or "", left_al)
            wc(9,  boat.crew_name or "", left_al)
            wc(10, boat.crew_name2 or "", left_al)
            wc(11, boat.crew_name3 or "", left_al)

            data_col = N_FIXED + 1
            for ri2, race in enumerate(races):
                c1r, c2r, c3r = data_col, data_col + 1, data_col + 2
                detail = race_result_map.get((race.id, boat.id))
                if detail is None:
                    for sc, pos in zip([c1r, c2r, c3r], ["left", "mid", "right"]):
                        c = ws.cell(row=row_num, column=sc, value="")
                        ap(c, font=norm9, alignment=center_al,
                           border=Border(left=med_side if pos == "left" else thin_side,
                                        right=med_side if pos == "right" else thin_side,
                                        top=top_bdr, bottom=thin_side))
                else:
                    fp   = detail["finish_position"]
                    code = detail["result_code"]
                    pts  = detail["points"]
                    # 着順: finish_position for OK, else code
                    if code == "OK":
                        disp_finish = fp if fp is not None else ""
                        disp_rank   = fp if fp is not None else ""
                    else:
                        disp_finish = fp if fp is not None else code
                        disp_rank   = code
                    for sc, val, pos in zip(
                        [c1r, c2r, c3r],
                        [disp_finish, disp_rank, pts if pts is not None else ""],
                        ["left", "mid", "right"]
                    ):
                        c = ws.cell(row=row_num, column=sc, value=val)
                        ap(c, font=norm9, alignment=center_al,
                           border=Border(left=med_side if pos == "left" else thin_side,
                                        right=med_side if pos == "right" else thin_side,
                                        top=top_bdr, bottom=thin_side))
                data_col += 3

            # Tail cells
            bdr_tail = Border(left=thin_side, right=thin_side, top=top_bdr, bottom=thin_side)
            if is_team:
                c = ws.cell(row=row_num, column=data_col, value=boat_total if boat_total is not None else "")
                ap(c, font=bold9, alignment=center_al, border=bdr_tail); data_col += 1
                c = ws.cell(row=row_num, column=data_col, value=team_total_val)
                ap(c, font=bold9 if team_total_val != "" else norm9, alignment=center_al, border=bdr_tail)
            else:
                c = ws.cell(row=row_num, column=data_col, value=boat_total if boat_total is not None else "")
                ap(c, font=bold9, alignment=center_al, border=bdr_tail)

        if not is_team:
            standings = calculate_individual_standings(tournament_id, db)
            # Filter to sheet's class if applicable
            if sheet_class:
                boat_ids_in_class = {b.id for b in boats}
                standings = [s for s in standings if s["boat_id"] in boat_ids_in_class]
            boat_map = {b.id: b for b in all_boats}
            for item in standings:
                boat = boat_map.get(item["boat_id"])
                if boat is None:
                    continue
                write_boat_row(current_row, item["rank"], boat.organization_name or "",
                               boat, item["net_points"], "", True,
                               boat_idx_in_team=0, team_boat_count=1)
                current_row += 1
        else:
            # Group boats by team
            team_boats_map: dict[str, list] = {}
            for boat in boats:
                team = boat.team_name or boat.organization_name or "未設定"
                team_boats_map.setdefault(team, []).append(boat)

            def boat_net(boat_id: int) -> int:
                return sum(
                    (race_result_map.get((r.id, boat_id)) or {}).get("points") or 0
                    for r in races
                )

            def team_net(tname: str) -> int:
                total = 0
                for race in races:
                    pts_list = sorted(
                        p for b in team_boats_map[tname]
                        if (p := (race_result_map.get((race.id, b.id)) or {}).get("points")) is not None
                    )
                    total += sum(pts_list[:team_size])
                return total

            def team_sort_key(tname: str) -> tuple:
                total = 0
                has_any = False
                is_incomplete = False
                for race in races:
                    pts_list = sorted(
                        p for b in team_boats_map[tname]
                        if (p := (race_result_map.get((race.id, b.id)) or {}).get("points")) is not None
                    )
                    if len(pts_list) > 0:
                        has_any = True
                        total += sum(pts_list[:team_size])
                        if len(pts_list) < team_size:
                            is_incomplete = True
                group = 0 if (has_any and not is_incomplete) else (1 if has_any else 2)
                return (group, total, tname)

            sorted_teams = sorted(team_boats_map.keys(), key=team_sort_key)
            n_total_teams = len(sorted_teams)
            for rank, tname in enumerate(sorted_teams, 1):
                t_total = team_net(tname)
                team_boat_list = team_boats_map[tname]
                team_start_row = current_row

                for i, boat in enumerate(team_boat_list):
                    write_boat_row(
                        current_row,
                        "",   # 順位: 結合セルで後設定
                        "",   # 大学名: 結合セルで後設定
                        boat,
                        boat_net(boat.id),
                        "",   # 大学合計: 結合セルで後設定
                        i == 0,
                        boat_idx_in_team=i,
                        team_boat_count=len(team_boat_list),
                    )
                    current_row += 1

                team_end_row = current_row - 1
                n_boats_in_team = len(team_boat_list)
                is_last_team = (rank == n_total_teams)
                # 大学合計列 = total_cols（最終列）
                for col_idx, val in [(1, rank), (3, tname), (total_cols, t_total)]:
                    if n_boats_in_team > 1:
                        ws.merge_cells(
                            start_row=team_start_row, start_column=col_idx,
                            end_row=team_end_row,     end_column=col_idx,
                        )
                    cell = ws.cell(row=team_start_row, column=col_idx, value=val)
                    cell.font      = bold9
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = Border(
                        top=med_side,
                        bottom=med_side if is_last_team else thin_side,
                        left=med_side   if col_idx == 1          else thin_side,
                        right=med_side  if col_idx == total_cols else thin_side,
                    )

        # ── 外枠太線: メインデータ表 ──────────────────────────────────────────
        apply_outer_border(ws, H1, current_row - 1, 1, total_cols)

        # ── Race info footer (得点表の下に1行空白を挟んで別表) ─────────────────
        footer_labels = ["レース日", "天気", "風向", "風速", "スタート",
                         "Top", "Last"]
        footer_start = current_row + 1   # current_row が空白行、+1 がフッター開始
        footer_end_row = footer_start + len(footer_labels) - 1
        footer_end_col = N_FIXED + n_races * 3   # 最終レース列まで

        # フッター全セルに細枠を先付け（中間空白列 2〜N_FIXED も含む）
        for fi in range(len(footer_labels)):
            r = footer_start + fi
            for ci2 in range(1, footer_end_col + 1):
                ws.cell(row=r, column=ci2).border = tb()

        # ラベル列 (col 1)
        for fi, label in enumerate(footer_labels):
            r = footer_start + fi
            c = ws.cell(row=r, column=1, value=label)
            ap(c, fill=gray_fill, font=bold9, alignment=left_al, border=tb())

        # レースデータ列 (得点表の R1, R2... 列と同じ位置)
        for ri, race in enumerate(races):
            base_col = N_FIXED + ri * 3 + 1
            vals = [
                getattr(race, "race_date",          "") or "",
                getattr(race, "weather",             "") or "",
                getattr(race, "wind_direction",      "") or "",
                getattr(race, "wind_speed",          "") or "",
                getattr(race, "start_time",          "") or "",
                getattr(race, "finish_time_top",     "") or "",
                getattr(race, "finish_time_last",    "") or "",
            ]
            for fi, val in enumerate(vals):
                r = footer_start + fi
                ws.merge_cells(start_row=r, start_column=base_col,
                               end_row=r, end_column=base_col + 2)
                c = ws.cell(row=r, column=base_col, value=val)
                ap(c, font=norm9, alignment=center_al, border=tb())

        # 外枠太線: レース情報表
        if n_races > 0:
            apply_outer_border(ws, footer_start, footer_end_row, 1, footer_end_col)

        # ── Column widths ─────────────────────────────────────────────────────
        col_widths = {1: 5, 2: 7, 3: 16, 4: 6, 5: 9,
                      6: 11, 7: 9, 8: 9, 9: 11, 10: 9, 11: 9}
        for ci, w in col_widths.items():
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ri2 in range(n_races):
            base = N_FIXED + ri2 * 3
            ws.column_dimensions[get_column_letter(base + 1)].width = 6
            ws.column_dimensions[get_column_letter(base + 2)].width = 6
            ws.column_dimensions[get_column_letter(base + 3)].width = 6
        tail_base = N_FIXED + n_races * 3 + 1
        ws.column_dimensions[get_column_letter(tail_base)].width = 8
        if is_team:
            ws.column_dimensions[get_column_letter(tail_base + 1)].width = 9

        # ── 空欄列の非表示 ────────────────────────────────────────────────────
        has_entry   = any(b.entry_number is not None for b in boats)
        has_boat_no = any(b.boat_number  for b in boats)
        has_helm2   = any(b.helmsman_name2 for b in boats)
        has_helm3   = any(b.helmsman_name3 for b in boats)
        has_crew2   = any(b.crew_name2   for b in boats)
        has_crew3   = any(b.crew_name3   for b in boats)
        if not has_entry:   ws.column_dimensions[get_column_letter(2)].hidden  = True
        if not has_boat_no: ws.column_dimensions[get_column_letter(4)].hidden  = True
        if not has_helm2:   ws.column_dimensions[get_column_letter(7)].hidden  = True
        if not has_helm3:   ws.column_dimensions[get_column_letter(8)].hidden  = True
        if not has_crew2:   ws.column_dimensions[get_column_letter(10)].hidden = True
        if not has_crew3:   ws.column_dimensions[get_column_letter(11)].hidden = True

        ws.freeze_panes = ws.cell(row=data_start, column=1)

    # ─── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    if len(classes) >= 2:
        for cls in classes:
            cls_boats = [b for b in all_boats if b.boat_class == cls]
            write_sheet(wb, cls, cls_boats, cls)

        # 総合 sheet: team totals across all classes
        ws_total = wb.create_sheet(title="総合")
        n_classes = len(classes)
        total_cols_ov = 2 + n_classes + 1  # 順位, 大学名, [class cols...], 合計

        date_str2 = date_str
        ws_total.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols_ov)
        c = ws_total.cell(row=1, column=1, value=f"{tournament.name}  総合")
        c.font = Font(bold=True, size=13)

        ov_hdrs = ["順位", "大学名"] + classes + ["合計得点"]
        for ci, h in enumerate(ov_hdrs, 1):
            c = ws_total.cell(row=3, column=ci, value=h)
            ap(c, fill=navy_fill, font=hdr_font, alignment=center_al, border=tb())
        ws_total.row_dimensions[3].height = 20

        # Compute team totals per class
        def class_team_total(tname: str, cls: str) -> int:
            cls_boats_t = [b for b in all_boats if b.boat_class == cls
                          and (b.team_name or b.organization_name) == tname]
            total = 0
            for race in races:
                pts_list = sorted(
                    p for b in cls_boats_t
                    if (p := (race_result_map.get((race.id, b.id)) or {}).get("points")) is not None
                )
                total += sum(pts_list[:team_size])
            return total

        # 全クラスに艇が登録されているチームのみ総合順位に含める
        team_classes: dict[str, set[str]] = {}
        for b in all_boats:
            tname = b.team_name or b.organization_name or "未設定"
            team_classes.setdefault(tname, set()).add(b.boat_class or "")
        all_team_names = sorted(
            t for t, clss in team_classes.items() if all(cls in clss for cls in classes)
        )

        team_grand = {t: sum(class_team_total(t, cls) for cls in classes) for t in all_team_names}

        def overall_sort_key(tname: str) -> tuple:
            grand = team_grand[tname]
            has_any = any(
                (race_result_map.get((r.id, b.id)) or {}).get("points") is not None
                for b in all_boats
                if (b.team_name or b.organization_name or "未設定") == tname
                for r in races
            )
            return (0 if has_any else 1, grand, tname)

        sorted_ov = sorted(all_team_names, key=overall_sort_key)

        for rank, tname in enumerate(sorted_ov, 1):
            row_num = rank + 3
            ws_total.cell(row=row_num, column=1, value=rank).border = tb()
            ws_total.cell(row=row_num, column=1).font = bold9
            ws_total.cell(row=row_num, column=1).alignment = center_al
            ws_total.cell(row=row_num, column=2, value=tname).border = tb()
            ws_total.cell(row=row_num, column=2).font = bold9
            for ci, cls in enumerate(classes, 3):
                ws_total.cell(row=row_num, column=ci, value=class_team_total(tname, cls)).border = tb()
                ws_total.cell(row=row_num, column=ci).alignment = center_al
                ws_total.cell(row=row_num, column=ci).font = norm9
            grand_col = 2 + n_classes + 1
            ws_total.cell(row=row_num, column=grand_col, value=team_grand[tname]).border = tb()
            ws_total.cell(row=row_num, column=grand_col).font = bold9
            ws_total.cell(row=row_num, column=grand_col).alignment = center_al

        ws_total.column_dimensions["A"].width = 5
        ws_total.column_dimensions["B"].width = 20
        for ci in range(3, 3 + n_classes + 1):
            ws_total.column_dimensions[get_column_letter(ci)].width = 12

    elif len(classes) == 1:
        cls_boats = [b for b in all_boats if b.boat_class == classes[0]]
        write_sheet(wb, classes[0], cls_boats, classes[0])
    else:
        write_sheet(wb, "成績", all_boats, None)

    return wb

def create_default_series_and_profiles(tournament, db):
    template = tournament.event_template

    created_series = []
    created_profiles = []

    def add_series(name, display_name, scheduled_races=None, max_races_per_day=None,
                   discard_type="NONE", discard_after_races=None, discard_count=0):
        s = Series(
            tournament_id=tournament.id,
            name=name,
            display_name=display_name,
            scheduled_races=scheduled_races,
            max_races_per_day=max_races_per_day,
            discard_type=discard_type,
            discard_after_races=discard_after_races,
            discard_count=discard_count,
        )
        db.add(s)
        db.flush()
        created_series.append(s)
        return s

    def add_profile(series, name, display_name, ranking_unit, class_scope,
                    scoring_team_size, school_score_method,
                    include_open_in_finish_scoring=True,
                    include_open_in_series_ranking=True,
                    include_open_in_school_ranking=False,
                    target_group_tag=None):
        p = RankingProfile(
            tournament_id=tournament.id,
            series_id=series.id if series else None,
            name=name,
            display_name=display_name,
            ranking_unit=ranking_unit,
            class_scope=class_scope,
            scoring_team_size=scoring_team_size,
            school_score_method=school_score_method,
            include_open_in_finish_scoring=include_open_in_finish_scoring,
            include_open_in_series_ranking=include_open_in_series_ranking,
            include_open_in_school_ranking=include_open_in_school_ranking,
            target_group_tag=target_group_tag,
        )
        db.add(p)
        created_profiles.append(p)
        return p

    if template == "INDIVIDUAL":
        main = add_series("main", "本戦", discard_type="WORST_N_AFTER_RACES", discard_after_races=5, discard_count=1)
        add_profile(main, "individual_470", "470個人順位", "boat", "470", 1, "INDIVIDUAL")
        add_profile(main, "individual_snipe", "スナイプ個人順位", "boat", "SNIPE", 1, "INDIVIDUAL")

    elif template == "TEAM_3_BOATS":
        main = add_series("main", "本戦", discard_type="NONE", discard_count=0)
        add_profile(main, "team_470", "470団体順位", "team", "470", 3, "SUM_TOP_N", False, False, False)
        add_profile(main, "team_snipe", "スナイプ団体順位", "team", "SNIPE", 3, "SUM_TOP_N", False, False, False)
        add_profile(main, "team_overall", "総合順位", "team", "ALL", 3, "ALL_RACES_SUM", False, False, False)

    elif template == "TEAM_4_BOATS":
        main = add_series("main", "本戦", discard_type="NONE", discard_count=0)
        add_profile(main, "team_470", "470団体順位", "team", "470", 4, "SUM_TOP_N", False, False, False)
        add_profile(main, "team_snipe", "スナイプ団体順位", "team", "SNIPE", 4, "SUM_TOP_N", False, False, False)
        add_profile(main, "team_overall", "総合順位", "team", "ALL", 4, "ALL_RACES_SUM", False, False, False)

    elif template == "WOMENS_BEST1_PER_CLASS":
        women = add_series("women", "女子レース", discard_type="WORST_N_AFTER_RACES", discard_after_races=5, discard_count=1)
        add_profile(women, "women_470_individual", "470個人順位", "boat", "470", 1, "INDIVIDUAL", True, True, False)
        add_profile(women, "women_snipe_individual", "スナイプ個人順位", "boat", "SNIPE", 1, "INDIVIDUAL", True, True, False)
        add_profile(women, "women_school_overall", "女子学校総合", "team", "ALL", 1, "BEST_ONE_EACH_CLASS", True, False, False)

    elif template == "MULTI_GROUP_HYBRID":
        main = add_series("main", "本戦", discard_type="NONE", discard_count=0)
        add_profile(main, "godai_team", "五大学戦順位", "team", "ALL", 3, "ALL_RACES_SUM", False, False, False, "godai")
        add_profile(main, "rokudai_team", "六大学戦順位", "team", "ALL", 3, "ALL_RACES_SUM", False, False, False, "rokudai")
        add_profile(main, "overall_team", "全体学校順位", "team", "ALL", 3, "ALL_RACES_SUM", False, False, False, "overall")
        add_profile(main, "overall_individual_470", "470全体個人順位", "boat", "470", 1, "INDIVIDUAL", True, True, False)
        add_profile(main, "overall_individual_snipe", "スナイプ全体個人順位", "boat", "SNIPE", 1, "INDIVIDUAL", True, True, False)

    db. commit()

def calculate_team_standings_by_class(
    tournament_id: int,
    db: Session,
    team_size: int,
    boat_class: str | None,
):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    boats_query = db.query(Boat).filter(Boat.tournament_id == tournament_id)

    if boat_class is not None and boat_class != "ALL":
        boats_query = boats_query.filter(Boat.boat_class == boat_class)

    boats = boats_query.order_by(Boat.id.asc()).all()

    races = (
        db.query(Race)
        .filter(Race.tournament_id == tournament_id)
        .order_by(Race.race_number.asc())
        .all()
    )

    if len(boats) == 0:
        return []

    boat_ids = [boat.id for boat in boats]

    results = (
        db.query(RaceResult)
        .filter(RaceResult.boat_id.in_(boat_ids))
        .join(Race, Race.id == RaceResult.race_id)
        .filter(Race.tournament_id == tournament_id)
        .all()
    )

    boat_map = {boat.id: boat for boat in boats}

    race_team_points = {race.id: {} for race in races}

    for result in results:
        if result.points is None:
            continue

        boat = boat_map.get(result.boat_id)
        if boat is None:
            continue

        team_name = boat.team_name if boat.team_name else boat.organization_name
        if not team_name:
            continue

        if team_name not in race_team_points[result.race_id]:
            race_team_points[result.race_id][team_name] = []

        race_team_points[result.race_id][team_name].append(result.points)

    all_team_names = set()
    for boat in boats:
        team_name = boat.team_name if boat.team_name else boat.organization_name
        if team_name:
            all_team_names.add(team_name)

    standings = []
    for team_name in sorted(all_team_names):
        race_points = []
        total_points = 0

        for race in races:
            points_list = sorted(race_team_points[race.id].get(team_name, []))
            if len(points_list) == 0:
                race_points.append(None)
            else:
                adopted = sum(points_list[:team_size])
                race_points.append(adopted)
                total_points += adopted

        is_incomplete = any(
            0 < len(race_team_points[race.id].get(team_name, [])) < team_size
            for race in races
        )
        has_any = any(p is not None for p in race_points)
        standings.append(
            {
                "boat_id": 0,
                "boat_number": "",
                "sail_number": "",
                "organization_name": team_name,
                "race_points": race_points,
                "total_points": total_points,
                "discarded_points": [],
                "net_points": total_points,
                "rank": 0,
                "_sort_group": 0 if (has_any and not is_incomplete) else (1 if has_any else 2),
            }
        )

    standings.sort(key=lambda x: (x["_sort_group"], x["net_points"], x["organization_name"]))
    for row in standings:
        del row["_sort_group"]

    for i, row in enumerate(standings, start=1):
        row["rank"] = i

    return standings

def calculate_team3_standings_sections(tournament_id: int, db: Session):
    return {
        "sections": [
            {
                "name": "470団体順位",
                "rows": calculate_team_standings_by_class(
                    tournament_id=tournament_id,
                    db=db,
                    team_size=3,
                    boat_class="470",
                ),
            },
            {
                "name": "スナイプ団体順位",
                "rows": calculate_team_standings_by_class(
                    tournament_id=tournament_id,
                    db=db,
                    team_size=3,
                    boat_class="SNIPE",
                ),
            },
            {
                "name": "総合順位",
                "rows": calculate_team_standings_by_class(
                    tournament_id=tournament_id,
                    db=db,
                    team_size=3,
                    boat_class=None,
                ),
            },
        ]
    }

def calculate_class_section_v3(
    tournament_id: int,
    db: Session,
    team_size: int,
    boat_class: str | None,
    section_title: str,
) -> dict:
    """
    指定クラスの帳票用セクションを構築する。
    返り値は ClassSection Pydantic モデルに対応する dict。
    """
    boats_query = db.query(Boat).filter(Boat.tournament_id == tournament_id)
    if boat_class is not None:
        boats_query = boats_query.filter(Boat.boat_class == boat_class)
    boats = boats_query.order_by(Boat.id.asc()).all()

    races = (
        db.query(Race)
        .filter(Race.tournament_id == tournament_id)
        .order_by(Race.race_number.asc())
        .all()
    )

    if len(boats) == 0:
        return {
            "class_name": boat_class or "ALL",
            "section_title": section_title,
            "race_count": len(races),
            "teams": [],
        }

    boat_ids = [boat.id for boat in boats]
    results = (
        db.query(RaceResult)
        .filter(RaceResult.boat_id.in_(boat_ids))
        .join(Race, Race.id == RaceResult.race_id)
        .filter(Race.tournament_id == tournament_id)
        .all()
    )
    result_map = {(r.race_id, r.boat_id): r for r in results}

    # カット設定を取得
    rule_config = db.query(RuleConfig).filter(RuleConfig.tournament_id == tournament_id).first()
    completed_races = len(races)
    discard_count = 0
    if (
        rule_config
        and rule_config.discard_enabled == 1
        and rule_config.discard_count is not None
        and rule_config.discard_count > 0
        and (
            rule_config.discard_start_race_count is None
            or completed_races >= rule_config.discard_start_race_count
        )
    ):
        discard_count = rule_config.discard_count

    team_cut_method = (rule_config.team_cut_method if rule_config and rule_config.team_cut_method else "individual")

    # チームごとに艇をグループ化
    team_boats: dict[str, list] = {}
    for boat in boats:
        tname = boat.team_name if boat.team_name else boat.organization_name
        if not tname:
            continue
        team_boats.setdefault(tname, []).append(boat)

    team_blocks = []
    for tname in sorted(team_boats.keys()):
        boats_in_team = team_boats[tname]

        # 艇別: 全レースの得点を収集（カット前）
        boat_rows = []
        for boat in boats_in_team:
            race_points = []
            for race in races:
                r = result_map.get((race.id, boat.id))
                pts = r.points if (r and r.points is not None) else None
                race_points.append(pts)

            boat_rows.append({
                "boat_id": boat.id,
                "sail_number": boat.sail_number,
                "helmsman_name": boat.helmsman_name,
                "crew_name": boat.crew_name,
                "race_points": race_points,
                "boat_total": 0,
                "discarded_race_indices": [],
            })

        # ---- カット処理 ----
        team_discarded_race_indices: list[int] = []

        if team_cut_method == "team" and discard_count > 0:
            # 方式A: チーム単位カット
            # レースごとの上位 team_size 艇合計を計算し、最大の discard_count レースをカット
            race_team_totals = []
            for race_idx in range(len(races)):
                per_race = sorted(
                    row["race_points"][race_idx]
                    for row in boat_rows
                    if row["race_points"][race_idx] is not None
                )
                race_team_totals.append(
                    sum(per_race[:team_size]) if per_race else None
                )
            valid = [(i, t) for i, t in enumerate(race_team_totals) if t is not None]
            worst = sorted(valid, key=lambda x: -x[1])[:discard_count]
            team_discarded_race_indices = [i for i, _ in worst]

            # 全艇に同じカットレースを適用
            for row in boat_rows:
                row["discarded_race_indices"] = team_discarded_race_indices
                raw = sum(p for p in row["race_points"] if p is not None)
                disc = sum(row["race_points"][i] for i in team_discarded_race_indices if row["race_points"][i] is not None)
                row["boat_total"] = raw - disc

        else:
            # 方式B: 個人単位カット（デフォルト）
            for row in boat_rows:
                raw_total = sum(p for p in row["race_points"] if p is not None)
                discarded_race_indices: list[int] = []
                if discard_count > 0:
                    valid_pts = [(i, p) for i, p in enumerate(row["race_points"]) if p is not None]
                    worst_pts = sorted(valid_pts, key=lambda x: -x[1])[:discard_count]
                    discarded_race_indices = [i for i, _ in worst_pts]
                disc_sum = sum(row["race_points"][i] for i in discarded_race_indices if row["race_points"][i] is not None)
                row["discarded_race_indices"] = discarded_race_indices
                row["boat_total"] = raw_total - disc_sum

        # レースごとのチーム合計（カット済み艇得点ベース、カットレースは None）
        team_race_totals = []
        team_total = 0
        for race_idx in range(len(races)):
            if race_idx in team_discarded_race_indices:
                team_race_totals.append(None)
            else:
                per_race = sorted(
                    row["race_points"][race_idx]
                    for row in boat_rows
                    if row["race_points"][race_idx] is not None
                )
                if len(per_race) == 0:
                    team_race_totals.append(None)
                else:
                    adopted = sum(per_race[:team_size])
                    team_race_totals.append(adopted)
                    team_total += adopted

        is_incomplete = any(
            0 < sum(1 for row in boat_rows if row["race_points"][i] is not None) < team_size
            for i in range(len(races))
            if i not in team_discarded_race_indices
            and any(row["race_points"][i] is not None for row in boat_rows)
        )
        has_any = any(t is not None for t in team_race_totals)
        team_blocks.append({
            "team_name": tname,
            "boats": boat_rows,
            "team_race_totals": team_race_totals,
            "team_total": team_total,
            "team_discarded_race_indices": team_discarded_race_indices,
            "rank": 0,
            "_sort_group": 0 if (has_any and not is_incomplete) else (1 if has_any else 2),
        })

    team_blocks.sort(key=lambda x: (x["_sort_group"], x["team_total"], x["team_name"]))
    for block in team_blocks:
        del block["_sort_group"]
    for i, block in enumerate(team_blocks, start=1):
        block["rank"] = i

    return {
        "class_name": boat_class or "ALL",
        "section_title": section_title,
        "race_count": len(races),
        "teams": team_blocks,
        "cut_method": team_cut_method,
    }


def build_overall_section_v3(class_sections: list[dict]) -> dict:
    """
    複数クラスのセクションから総合順位セクションを構築する。
    """
    class_names = [s["class_name"] for s in class_sections]

    team_scores: dict[str, dict[str, int]] = {}
    for section in class_sections:
        cls_name = section["class_name"]
        for block in section["teams"]:
            tname = block["team_name"]
            team_scores.setdefault(tname, {})
            team_scores[tname][cls_name] = block["team_total"]

    overall_teams = []
    for tname in sorted(team_scores.keys()):
        # 全クラスに出場しているチームのみ総合順位に含める
        if len(team_scores[tname]) < len(class_names):
            continue
        class_scores = [
            {"class_name": cn, "points": team_scores[tname].get(cn, 0)}
            for cn in class_names
        ]
        total = sum(item["points"] for item in class_scores)
        overall_teams.append({
            "team_name": tname,
            "class_scores": class_scores,
            "total_points": total,
            "rank": 0,
        })

    overall_teams.sort(key=lambda x: (x["total_points"], x["team_name"]))
    for i, row in enumerate(overall_teams, start=1):
        row["rank"] = i

    return {
        "section_title": "総合順位",
        "teams": overall_teams,
    }


@app.get("/")
def read_root():
    return {"message": "Hello Sailing App"}

@app.get("/users/me", response_model=UserOut)
def get_current_user_info(current_user=Depends(get_current_user)):
    return current_user


@app.get("/tournaments", response_model=list[TournamentOut])
def get_tournaments(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role == "admin":
        return db.query(Tournament).filter(Tournament.deleted_at == None).all()
    ids = [
        m.tournament_id
        for m in db.query(TournamentMember)
            .filter(TournamentMember.user_id == current_user.id)
            .all()
    ]
    return db.query(Tournament).filter(
        Tournament.id.in_(ids),
        Tournament.deleted_at == None,
    ).all()


@app.get("/tournaments/trash", response_model=list[TournamentOut])
def get_trash(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role == "admin":
        return db.query(Tournament).filter(Tournament.deleted_at != None).all()
    owner_ids = [
        m.tournament_id
        for m in db.query(TournamentMember)
            .filter(TournamentMember.user_id == current_user.id, TournamentMember.role == "owner")
            .all()
    ]
    return db.query(Tournament).filter(
        Tournament.id.in_(owner_ids),
        Tournament.deleted_at != None,
    ).all()

@app.post("/tournaments", response_model=TournamentOut)
def create_tournament(
    tournament: TournamentCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    new_tournament = Tournament(**tournament.model_dump(), owner_id=current_user.id)
    db.add(new_tournament)
    db.flush()  # id を確定させてから tournament_members に追加

    # 作成者を owner として登録
    db.add(TournamentMember(
        tournament_id=new_tournament.id,
        user_id=current_user.id,
        role="owner",
    ))

    rule_config = RuleConfig(
        tournament_id=new_tournament.id,
        scheduled_races=1,
        minimum_races_for_series=1,
        discard_enabled=0,
        discard_start_race_count=None,
        discard_count=None,
        dnc_rule="ENTRIES_PLUS_1",
        dns_rule="ENTRIES_PLUS_1",
        ocs_rule="STARTERS_PLUS_1",
        dnf_rule="STARTERS_PLUS_1",
        ret_rule="STARTERS_PLUS_1",
        dsq_rule="STARTERS_PLUS_1",
        ufd_rule="ENTRIES_PLUS_1",
        bfd_rule="ENTRIES_PLUS_1",
    )
    db.add(rule_config)

    create_default_series_and_profiles(new_tournament, db)

    db.commit()
    db.refresh(new_tournament)
    return new_tournament


@app.get("/tournaments/{tournament_id}/series", response_model=list[SeriesOut])
def get_series(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db)
    return (
        db.query(Series)
        .filter(Series.tournament_id == tournament_id)
        .order_by(Series.id.asc())
        .all()
    )


@app.get("/tournaments/{tournament_id}/ranking-profiles", response_model=list[RankingProfileOut])
def get_ranking_profiles(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db)
    return (
        db.query(RankingProfile)
        .filter(RankingProfile.tournament_id == tournament_id)
        .order_by(RankingProfile.id.asc())
        .all()
    )

@app.get("/tournaments/{tournament_id}/members", response_model=list[TournamentMemberOut])
def list_tournament_members(
    tournament_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """大会メンバー一覧（owner/editor 本人または admin のみ）"""
    check_tournament_access(tournament_id, current_user, db)
    members = (
        db.query(TournamentMember)
        .filter(TournamentMember.tournament_id == tournament_id)
        .all()
    )
    result = []
    for m in members:
        u = db.query(User).filter(User.id == m.user_id).first()
        result.append(TournamentMemberOut(
            user_id=m.user_id,
            email=u.email if u else "unknown",
            role=m.role,
        ))
    return result


@app.post("/tournaments/{tournament_id}/members", response_model=TournamentMemberOut)
def add_tournament_editor(
    tournament_id: int,
    body: AddMemberRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """メールアドレスで editor を追加（owner または admin のみ）"""
    check_tournament_access(tournament_id, current_user, db, owner_only=True)

    # ローカル users テーブルを検索
    target = db.query(User).filter(User.email == body.email).first()

    # ローカルに存在しない場合、Supabase auth を検索して自動登録
    # （サインアップ済みだが未ログインのユーザーへの対応）
    if target is None and AUTH_ENABLED:
        try:
            supabase = get_supabase()
            resp = supabase.auth.admin.list_users()
            users_list = resp if isinstance(resp, list) else getattr(resp, "users", [])
            sb_user = next(
                (u for u in users_list if (u.email or "").lower() == body.email.lower()),
                None,
            )
            if sb_user:
                target = User(id=sb_user.id, email=sb_user.email or body.email, role="member")
                db.add(target)
                db.flush()
        except Exception as e:
            print(f"[add_member] Supabase lookup error: {e}", flush=True)

    if target is None:
        raise HTTPException(
            status_code=404,
            detail=f"{body.email} のユーザーが見つかりません。先にサインアップしてください。",
        )

    existing = db.query(TournamentMember).filter(
        TournamentMember.tournament_id == tournament_id,
        TournamentMember.user_id == target.id,
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="既にメンバーとして登録されています")

    db.add(TournamentMember(tournament_id=tournament_id, user_id=target.id, role="editor"))
    db.commit()
    return TournamentMemberOut(user_id=target.id, email=target.email, role="editor")


@app.get("/tournaments/{tournament_id}/boats", response_model=list[BoatOut])
def get_boats(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db)
    return db.query(Boat).filter(Boat.tournament_id == tournament_id).all()


@app.post("/tournaments/{tournament_id}/boats/import")
async def import_boats_csv(
    tournament_id: int,
    file: UploadFile = File(...),
    boat_class: str | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """CSVファイルから艇を一括登録。sail_number が重複する行はスキップ。"""
    check_tournament_access(tournament_id, current_user, db)

    content = await file.read()
    for _enc in ("utf-8-sig", "cp932", "shift-jis", "utf-8"):
        try:
            text_data = content.decode(_enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        text_data = content.decode("utf-8", errors="replace")

    reader = csv_module.DictReader(io.StringIO(text_data))

    imported = 0
    skipped = 0

    def _s(v: str | None) -> str | None:
        return v.strip() or None if v else None

    for row in reader:
        sail_number = (row.get("sail_number") or "").strip()
        if not sail_number:
            skipped += 1
            continue

        entry_raw = (row.get("entry_number") or "").strip()
        entry_num = int(entry_raw) if entry_raw.isdigit() else None
        resolved_class = boat_class or _s(row.get("boat_class"))

        existing = db.query(Boat).filter(
            Boat.tournament_id == tournament_id,
            Boat.sail_number == sail_number,
        ).first()
        if existing:
            # sail_number が一致する艇は上書き更新
            existing.entry_number     = entry_num
            existing.boat_number      = _s(row.get("boat_number"))
            existing.organization_name= _s(row.get("organization_name"))
            existing.helmsman_name    = _s(row.get("helmsman_name"))
            existing.helmsman_name2   = _s(row.get("helmsman_name2"))
            existing.helmsman_name3   = _s(row.get("helmsman_name3"))
            existing.crew_name        = _s(row.get("crew_name"))
            existing.crew_name2       = _s(row.get("crew_name2"))
            existing.crew_name3       = _s(row.get("crew_name3"))
            if resolved_class:
                existing.boat_class   = resolved_class
            if _s(row.get("team_name")):
                existing.team_name    = _s(row.get("team_name"))
            imported += 1
            continue

        db.add(Boat(
            tournament_id=tournament_id,
            entry_number=entry_num,
            boat_number=_s(row.get("boat_number")),
            sail_number=sail_number,
            organization_name=_s(row.get("organization_name")),
            helmsman_name=_s(row.get("helmsman_name")),
            helmsman_name2=_s(row.get("helmsman_name2")),
            helmsman_name3=_s(row.get("helmsman_name3")),
            crew_name=_s(row.get("crew_name")),
            crew_name2=_s(row.get("crew_name2")),
            crew_name3=_s(row.get("crew_name3")),
            boat_class=resolved_class,
            team_name=_s(row.get("team_name")),
        ))
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped}


@app.post("/tournaments/{tournament_id}/boats", response_model=BoatOut)
def create_boat(
    tournament_id: int,
    boat: BoatCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_tournament_access(tournament_id, current_user, db)
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    sail = (boat.sail_number or "").strip()
    if sail:
        dup = db.query(Boat).filter(Boat.tournament_id == tournament_id, Boat.sail_number == sail).first()
        if dup:
            raise HTTPException(status_code=409, detail=f"セールNo. '{sail}' は既に登録されています")
    if boat.entry_number is not None:
        dup = db.query(Boat).filter(Boat.tournament_id == tournament_id, Boat.entry_number == boat.entry_number).first()
        if dup:
            raise HTTPException(status_code=409, detail=f"Entry No. {boat.entry_number} は既に登録されています")

    data = boat.model_dump()
    data["sail_number"] = sail  # None → ""
    new_boat = Boat(tournament_id=tournament_id, **data)
    db.add(new_boat)
    db.commit()
    db.refresh(new_boat)
    return new_boat


@app.put("/boats/{boat_id}", response_model=BoatOut)
def update_boat(
    boat_id: int,
    boat: BoatCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    existing = db.query(Boat).filter(Boat.id == boat_id).first()
    if existing is None:
        raise HTTPException(status_code=404, detail="Boat not found")
    check_tournament_access(existing.tournament_id, current_user, db)

    sail = (boat.sail_number or "").strip()
    if sail:
        dup = db.query(Boat).filter(
            Boat.tournament_id == existing.tournament_id,
            Boat.sail_number == sail,
            Boat.id != boat_id,
        ).first()
        if dup:
            raise HTTPException(status_code=409, detail=f"セールNo. '{sail}' は既に登録されています")
    if boat.entry_number is not None:
        dup = db.query(Boat).filter(
            Boat.tournament_id == existing.tournament_id,
            Boat.entry_number == boat.entry_number,
            Boat.id != boat_id,
        ).first()
        if dup:
            raise HTTPException(status_code=409, detail=f"Entry No. {boat.entry_number} は既に登録されています")

    data = boat.model_dump()
    data["sail_number"] = sail
    for field, value in data.items():
        setattr(existing, field, value)
    db.commit()
    db.refresh(existing)
    return existing


@app.delete("/boats/{boat_id}", status_code=204)
def delete_boat(
    boat_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    existing = db.query(Boat).filter(Boat.id == boat_id).first()
    if existing is None:
        raise HTTPException(status_code=404, detail="Boat not found")
    check_tournament_access(existing.tournament_id, current_user, db)
    # FK 制約を回避するため関連するレース結果を先に削除
    db.query(RaceResult).filter(RaceResult.boat_id == boat_id).delete()
    db.delete(existing)
    db.commit()


@app.patch("/tournaments/{tournament_id}/boats/bulk")
def bulk_update_boats(
    tournament_id: int,
    payload: BoatBulkUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """艇を一括更新: id あり→更新, id なし→新規追加, deleted_ids→削除"""
    check_tournament_access(tournament_id, current_user, db)
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    is_team = tournament.event_template in ("TEAM_3_BOATS", "TEAM_4_BOATS", "MULTI_GROUP_HYBRID")

    # 削除
    for bid in payload.deleted_ids:
        boat = db.query(Boat).filter(Boat.id == bid, Boat.tournament_id == tournament_id).first()
        if boat:
            db.query(RaceResult).filter(RaceResult.boat_id == bid).delete()
            db.delete(boat)

    updated = 0
    created = 0
    for item in payload.boats:
        sail = (item.sail_number or "").strip() or None
        if sail is None and item.entry_number is None:
            continue  # 識別子なし行はスキップ

        team_name = item.team_name
        if is_team and not team_name:
            team_name = (item.organization_name or "").strip() or None

        if item.id is not None:
            boat = db.query(Boat).filter(
                Boat.id == item.id, Boat.tournament_id == tournament_id
            ).first()
            if boat:
                boat.entry_number      = item.entry_number
                boat.boat_number       = (item.boat_number or "").strip() or None
                boat.sail_number       = sail or boat.sail_number
                boat.organization_name = (item.organization_name or "").strip() or None
                boat.helmsman_name     = (item.helmsman_name  or "").strip() or None
                boat.helmsman_name2    = (item.helmsman_name2 or "").strip() or None
                boat.helmsman_name3    = (item.helmsman_name3 or "").strip() or None
                boat.crew_name         = (item.crew_name  or "").strip() or None
                boat.crew_name2        = (item.crew_name2 or "").strip() or None
                boat.crew_name3        = (item.crew_name3 or "").strip() or None
                if item.boat_class:
                    boat.boat_class    = item.boat_class
                boat.team_name         = team_name
                updated += 1
        else:
            db.add(Boat(
                tournament_id=tournament_id,
                entry_number=item.entry_number,
                boat_number=(item.boat_number or "").strip() or None,
                sail_number=sail,
                organization_name=(item.organization_name or "").strip() or None,
                helmsman_name=(item.helmsman_name  or "").strip() or None,
                helmsman_name2=(item.helmsman_name2 or "").strip() or None,
                helmsman_name3=(item.helmsman_name3 or "").strip() or None,
                crew_name=(item.crew_name  or "").strip() or None,
                crew_name2=(item.crew_name2 or "").strip() or None,
                crew_name3=(item.crew_name3 or "").strip() or None,
                boat_class=item.boat_class,
                team_name=team_name,
            ))
            created += 1

    db.commit()
    return {"updated": updated, "created": created, "deleted": len(payload.deleted_ids)}


@app.get("/tournaments/{tournament_id}/rules", response_model=RuleConfigOut)
def get_rule_config(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db)

    rule_config = db.query(RuleConfig).filter(RuleConfig.tournament_id == tournament_id).first()

    if rule_config is None:
        rule_config = RuleConfig(
            tournament_id=tournament_id,
            scheduled_races=1,
            minimum_races_for_series=1,
            discard_enabled=0,
            discard_start_race_count=None,
            discard_count=None,
            dnc_rule="ENTRIES_PLUS_1",
            dns_rule="ENTRIES_PLUS_1",
            ocs_rule="STARTERS_PLUS_1",
            dnf_rule="STARTERS_PLUS_1",
            ret_rule="STARTERS_PLUS_1",
            dsq_rule="STARTERS_PLUS_1",
            ufd_rule="ENTRIES_PLUS_1",
            bfd_rule="ENTRIES_PLUS_1",
        )
        db.add(rule_config)
        db.commit()
        db.refresh(rule_config)

    return rule_config


@app.put("/tournaments/{tournament_id}/rules", response_model=RuleConfigOut)
def update_rule_config(
    tournament_id: int,
    payload: RuleConfigUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_tournament_access(tournament_id, current_user, db)
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    rule_config = db.query(RuleConfig).filter(RuleConfig.tournament_id == tournament_id).first()

    if rule_config is None:
        rule_config = RuleConfig(tournament_id=tournament_id)
        db.add(rule_config)

    data = payload.model_dump()

    for key, value in data.items():
        if key == "discard_enabled":
            setattr(rule_config, key, 1 if value else 0)
        else:
            setattr(rule_config, key, value)

    db.commit()
    db.refresh(rule_config)
    return rule_config

@app.get("/tournaments/{tournament_id}/races", response_model=list[RaceOut])
def get_races(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db)
    races = (
        db.query(Race)
        .filter(Race.tournament_id == tournament_id)
        .order_by(Race.race_number.asc())
        .all()
    )
    return races


@app.post("/tournaments/{tournament_id}/races", response_model=RaceOut)
def create_race(
    tournament_id: int,
    race: RaceCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_tournament_access(tournament_id, current_user, db)
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    existing = (
        db.query(Race)
        .filter(Race.tournament_id == tournament_id, Race.race_number == race.race_number)
        .first()
    )
    if existing is not None:
        raise HTTPException(status_code=400, detail="Race number already exists")

    new_race = Race(
        tournament_id=tournament_id,
        race_number=race.race_number,
        name=race.name,
        status=race.status,
    )
    db.add(new_race)
    db.commit()
    db.refresh(new_race)
    return new_race

@app.put("/races/{race_id}", response_model=RaceOut)
def update_race(
    race_id: int,
    body: RaceUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    race = db.query(Race).filter(Race.id == race_id).first()
    if race is None:
        raise HTTPException(status_code=404, detail="Race not found")
    check_tournament_access(race.tournament_id, current_user, db)
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(race, field, value)
    db.commit()
    db.refresh(race)
    return race

@app.get("/races/{race_id}/results", response_model=list[RaceResultOut])
def get_race_results(race_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    race = db.query(Race).filter(Race.id == race_id).first()
    if race is None:
        raise HTTPException(status_code=404, detail="Race not found")
    check_tournament_access(race.tournament_id, current_user, db)
    results = (
        db.query(RaceResult)
        .filter(RaceResult.race_id == race_id)
        .order_by(RaceResult.id.asc())
        .all()
    )
    return results

@app.put("/races/{race_id}/results", response_model=list[RaceResultOut])
def save_race_results(
    race_id: int,
    payload: list[RaceResultInput],
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    race = db.query(Race).filter(Race.id == race_id).first()
    if race is None:
        raise HTTPException(status_code=404, detail="Race not found")

    tournament = db.query(Tournament).filter(Tournament.id == race.tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    check_tournament_access(tournament.id, current_user, db)

    rule_config = db.query(RuleConfig).filter(RuleConfig.tournament_id == tournament.id).first()
    if rule_config is None:
        raise HTTPException(status_code=404, detail="RuleConfig not found")

    # Duplicate finish_position check per class (each class has independent positions)
    boat_ids_in_payload = [item.boat_id for item in payload]
    boat_class_map = {
        b.id: (b.boat_class or "")
        for b in db.query(Boat).filter(Boat.id.in_(boat_ids_in_payload)).all()
    }
    class_pos_seen: dict[str, set] = {}
    for item in payload:
        if item.finish_position is not None:
            cls = boat_class_map.get(item.boat_id, "")
            seen = class_pos_seen.setdefault(cls, set())
            if item.finish_position in seen:
                raise HTTPException(status_code=400, detail="Duplicate finish_position detected")
            seen.add(item.finish_position)

    # クラス別にエントリー数・スターター数を計算
    all_classes = set(boat_class_map.values())
    class_entries: dict[str, int] = {
        cls: get_entries_count(tournament.id, db, boat_class=cls or None)
        for cls in all_classes
    }
    class_starters: dict[str, int] = {
        cls: get_starters_count(payload, boat_class_map, cls)
        for cls in all_classes
    }

    db.query(RaceResult).filter(RaceResult.race_id == race_id).delete()

    new_results = []
    for item in payload:
        cls = boat_class_map.get(item.boat_id, "")
        points = calculate_points_for_result(
            item=item,
            rule_config=rule_config,
            entries_count=class_entries.get(cls, 0),
            starters_count=class_starters.get(cls, 0),
        )

        new_result = RaceResult(
            race_id=race_id,
            boat_id=item.boat_id,
            finish_position=item.finish_position,
            result_code=item.result_code,
            points=points,
            note=item.note,
        )
        db.add(new_result)
        new_results.append(new_result)

    db.commit()

    for result in new_results:
        db.refresh(result)

    return new_results

@app.get("/tournaments/{tournament_id}/standings", response_model=list[StandingRow])
def get_standings(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    check_tournament_access(tournament_id, current_user, db)
    return calculate_standings(tournament_id, db)

@app.get("/tournaments/{tournament_id}/standings-v2", response_model=StandingsResponse)
def get_standings_v2(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db)

    if tournament.event_template == "TEAM_3_BOATS":
        return calculate_team3_standings_sections(tournament_id, db)

    if tournament.event_template == "INDIVIDUAL":
        return {
            "sections": [
                {
                    "name": "総合順位",
                    "rows": calculate_individual_standings(tournament_id, db),
                }
            ]
        }

    if tournament.event_template == "TEAM_4_BOATS":
        return {
            "sections": [
                {
                    "name": "総合順位",
                    "rows": calculate_team_standings_by_class(
                        tournament_id=tournament_id,
                        db=db,
                        team_size=4,
                        boat_class=None,
                    ),
                }
            ]
        }

    raise HTTPException(status_code=501, detail="standings-v2 not implemented for this event_template yet")

@app.get("/tournaments/{tournament_id}/standings-v3", response_model=StandingsV3Response)
def get_standings_v3(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db)

    if tournament.event_template != "TEAM_3_BOATS":
        raise HTTPException(status_code=501, detail="standings-v3 is only implemented for TEAM_3_BOATS")

    team_size = 3
    classes = parse_class_config(tournament.class_config)

    class_sections = []
    if classes:
        for entry in classes:
            display = class_display_name(entry)
            section = calculate_class_section_v3(
                tournament_id=tournament_id,
                db=db,
                team_size=team_size,
                boat_class=display,
                section_title=f"{display}団体順位",
            )
            class_sections.append(section)
    else:
        # class_config 未設定時は全艇を1セクションとして扱う
        section = calculate_class_section_v3(
            tournament_id=tournament_id,
            db=db,
            team_size=team_size,
            boat_class=None,
            section_title="団体順位",
        )
        class_sections.append(section)

    overall_section = None
    if len(class_sections) > 1:
        overall_section = build_overall_section_v3(class_sections)

    return {
        "event_template": tournament.event_template,
        "class_sections": class_sections,
        "overall_section": overall_section,
    }

@app.get("/tournaments/{tournament_id}/export/excel")
def export_excel(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    check_tournament_access(tournament_id, current_user, db)
    wb = build_standings_workbook(tournament_id, db)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    filename = f"tournament_{tournament_id}_standings.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )

@app.get("/tournaments/{tournament_id}/export/pdf")
def export_pdf(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.lib.units import mm
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab がインストールされていません")

    check_tournament_access(tournament_id, current_user, db)
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
        font_name = "HeiseiKakuGo-W5"
    except Exception:
        font_name = "Helvetica"

    is_team   = tournament.event_template in ("TEAM_3_BOATS", "TEAM_4_BOATS", "MULTI_GROUP_HYBRID")
    team_size = 4 if tournament.event_template == "TEAM_4_BOATS" else 3

    all_boats = (
        db.query(Boat)
        .filter(Boat.tournament_id == tournament_id)
        .order_by(Boat.id)
        .all()
    )
    races, race_result_map = get_race_result_details_by_boat(tournament_id, db)
    classes = _parse_class_config(tournament.class_config)
    now = datetime.now()
    navy = colors.HexColor("#1F4E78")

    def _normal(size=8):
        return ParagraphStyle("n", fontName=font_name, fontSize=size)

    def build_elements(boats_list, sheet_class):
        elems = []
        title = tournament.name + (f"  {sheet_class}" if sheet_class else "")
        elems.append(Paragraph(title, ParagraphStyle("t", fontName=font_name, fontSize=13, spaceAfter=2)))
        ts = f"更新日時: {now.strftime('%Y年%m月%d日 %H:%M')}"
        elems.append(Paragraph(ts, ParagraphStyle("ts", fontName=font_name, fontSize=7,
                                                  textColor=colors.HexColor("#888888"), spaceAfter=4)))

        n_races = len(races)
        fixed_hdr = ["順位", "大学名", "セールNo.", "スキッパー", "クルー"]
        race_hdrs = [f"R{r.race_number}" for r in races]
        tail_hdrs = ["艇計", "大学計"] if is_team else ["合計"]
        headers = [fixed_hdr + race_hdrs + tail_hdrs]

        def _pts(bid, race):
            d = race_result_map.get((race.id, bid))
            return str(d["points"]) if d and d.get("points") is not None else ""

        rows_data = []
        if not is_team:
            standings = calculate_individual_standings(tournament_id, db)
            bmap = {b.id: b for b in all_boats}
            bid_set = {b.id for b in boats_list} if sheet_class else None
            for item in standings:
                if bid_set and item["boat_id"] not in bid_set:
                    continue
                b = bmap.get(item["boat_id"])
                if not b:
                    continue
                rows_data.append([
                    str(item["rank"]), b.organization_name or "", b.sail_number or "",
                    b.helmsman_name or "", b.crew_name or "",
                ] + [_pts(b.id, r) for r in races] + [str(item["net_points"])])
        else:
            tbm: dict[str, list] = {}
            for b in boats_list:
                t = b.team_name or b.organization_name or "未設定"
                tbm.setdefault(t, []).append(b)

            def _tnet(tn):
                tot = 0
                for race in races:
                    pl = sorted(p for b in tbm[tn]
                                if (p := (race_result_map.get((race.id, b.id)) or {}).get("points")) is not None)
                    tot += sum(pl[:team_size])
                return tot

            def _bnet(bid):
                return sum((race_result_map.get((r.id, bid)) or {}).get("points") or 0 for r in races)

            for rank, tn in enumerate(sorted(tbm, key=lambda t: (_tnet(t), t)), 1):
                tt = _tnet(tn)
                for i, b in enumerate(tbm[tn]):
                    rows_data.append([
                        str(rank) if i == 0 else "",
                        tn if i == 0 else "",
                        b.sail_number or "",
                        b.helmsman_name or "",
                        b.crew_name or "",
                    ] + [_pts(b.id, r) for r in races]
                    + [str(_bnet(b.id)), str(tt) if i == 0 else ""])

        cw = [10*mm, 32*mm, 18*mm, 26*mm, 26*mm] + [14*mm]*n_races + ([14*mm, 14*mm] if is_team else [14*mm])
        tbl = Table(headers + rows_data, colWidths=cw, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  navy),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, -1), font_name),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("ALIGN",         (1, 1), (1, -1),  "LEFT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.black),
            ("BOX",           (0, 0), (-1, -1), 1.2, colors.black),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        elems.append(tbl)

        if races:
            elems.append(Spacer(1, 5*mm))
            flabels = ["レース日", "天気", "風向", "風速", "スタート", "Top", "Last"]
            fattrs  = ["race_date","weather","wind_direction","wind_speed",
                       "start_time","finish_time_top","finish_time_last"]
            fdata = [[""] + [f"R{r.race_number}" for r in races]]
            for lbl, attr in zip(flabels, fattrs):
                fdata.append([lbl] + [getattr(r, attr) or "" for r in races])
            fcw = [25*mm] + [14*mm]*n_races
            ftbl = Table(fdata, colWidths=fcw)
            ftbl.setStyle(TableStyle([
                ("FONTNAME",  (0, 0), (-1, -1), font_name),
                ("FONTSIZE",  (0, 0), (-1, -1), 8),
                ("BACKGROUND",(0, 0), (-1, 0),  colors.HexColor("#F0F4F8")),
                ("GRID",      (0, 0), (-1, -1), 0.4, colors.black),
                ("BOX",       (0, 0), (-1, -1), 1.2, colors.black),
                ("ALIGN",     (0, 0), (-1, -1), "CENTER"),
                ("ALIGN",     (0, 1), (0, -1),  "LEFT"),
                ("VALIGN",    (0, 0), (-1, -1), "MIDDLE"),
            ]))
            elems.append(ftbl)

        return elems

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    all_elems = []
    if len(classes) >= 2:
        for cls in classes:
            all_elems.extend(build_elements([b for b in all_boats if b.boat_class == cls], cls))
    elif len(classes) == 1:
        all_elems.extend(build_elements([b for b in all_boats if b.boat_class == classes[0]], classes[0]))
    else:
        all_elems.extend(build_elements(all_boats, None))

    doc.build(all_elems)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="tournament_{tournament_id}_standings.pdf"'},
    )


@app.post("/tournaments/{tournament_id}/seed/team3-demo")
def seed_team3_demo(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    check_tournament_access(tournament_id, current_user, db, owner_only=True)
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")

    if tournament.event_template != "TEAM_3_BOATS":
        raise HTTPException(status_code=400, detail="This seed is only for TEAM_3_BOATS tournaments")

    rule_config = db.query(RuleConfig).filter(RuleConfig.tournament_id == tournament_id).first()
    if rule_config is None:
        rule_config = RuleConfig(
            tournament_id=tournament_id,
            scheduled_races=1,
            minimum_races_for_series=1,
            discard_enabled=0,
            discard_start_race_count=None,
            discard_count=None,
            dnc_rule="ENTRIES_PLUS_1",
            dns_rule="ENTRIES_PLUS_1",
            ocs_rule="STARTERS_PLUS_1",
            dnf_rule="STARTERS_PLUS_1",
            ret_rule="STARTERS_PLUS_1",
            dsq_rule="STARTERS_PLUS_1",
            ufd_rule="ENTRIES_PLUS_1",
            bfd_rule="ENTRIES_PLUS_1",
        )
        db.add(rule_config)
        db.flush()

    existing_boats = db.query(Boat).filter(Boat.tournament_id == tournament_id).count()
    if existing_boats == 0:
        demo_boats = [
            Boat(tournament_id=tournament_id, boat_number="1", sail_number="A1", organization_name="A大学", team_name="A大学", boat_class="470", helmsman_name="A1", crew_name="A1c"),
            Boat(tournament_id=tournament_id, boat_number="2", sail_number="A2", organization_name="A大学", team_name="A大学", boat_class="470", helmsman_name="A2", crew_name="A2c"),
            Boat(tournament_id=tournament_id, boat_number="3", sail_number="A3", organization_name="A大学", team_name="A大学", boat_class="470", helmsman_name="A3", crew_name="A3c"),
            Boat(tournament_id=tournament_id, boat_number="4", sail_number="B1", organization_name="B大学", team_name="B大学", boat_class="470", helmsman_name="B1", crew_name="B1c"),
            Boat(tournament_id=tournament_id, boat_number="5", sail_number="B2", organization_name="B大学", team_name="B大学", boat_class="470", helmsman_name="B2", crew_name="B2c"),
            Boat(tournament_id=tournament_id, boat_number="6", sail_number="B3", organization_name="B大学", team_name="B大学", boat_class="470", helmsman_name="B3", crew_name="B3c"),
        ]
        db.add_all(demo_boats)
        db.flush()

    boats = (
        db.query(Boat)
        .filter(Boat.tournament_id == tournament_id)
        .order_by(Boat.id.asc())
        .all()
    )

    race = (
        db.query(Race)
        .filter(Race.tournament_id == tournament_id, Race.race_number == 1)
        .first()
    )
    if race is None:
        race = Race(
            tournament_id=tournament_id,
            race_number=1,
            name="Race 1",
            status="CONFIRMED",
        )
        db.add(race)
        db.flush()

    existing_results = db.query(RaceResult).filter(RaceResult.race_id == race.id).count()
    if existing_results == 0:
        for i, boat in enumerate(boats[:6], start=1):
            db.add(
                RaceResult(
                    race_id=race.id,
                    boat_id=boat.id,
                    finish_position=i,
                    result_code="OK",
                    points=i,
                    note=None,
                )
            )

    db.commit()

    return {
        "message": "team3 demo seeded",
        "tournament_id": tournament_id,
        "race_id": race.id,
        "boat_count": len(boats),
    }


# ─── /tournaments/{id} 単体（全サブリソースルートより後に登録） ───────────────

@app.get("/tournaments/{tournament_id}", response_model=TournamentOut)
def get_tournament(tournament_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db)
    return tournament


@app.put("/tournaments/{tournament_id}", response_model=TournamentOut)
def update_tournament(
    tournament_id: int,
    body: TournamentUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db, owner_only=True)

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(tournament, field, value)

    db.commit()
    db.refresh(tournament)
    return tournament


@app.delete("/tournaments/{tournament_id}", status_code=204)
def trash_tournament(
    tournament_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """ゴミ箱に移動（ソフトデリート）"""
    tournament = db.query(Tournament).filter(
        Tournament.id == tournament_id, Tournament.deleted_at == None
    ).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found")
    check_tournament_access(tournament_id, current_user, db, owner_only=True)
    tournament.deleted_at = datetime.now(timezone.utc).isoformat()
    db.commit()


@app.post("/tournaments/{tournament_id}/restore", status_code=204)
def restore_tournament(
    tournament_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """ゴミ箱から復元"""
    tournament = db.query(Tournament).filter(
        Tournament.id == tournament_id, Tournament.deleted_at != None
    ).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found in trash")
    check_tournament_access(tournament_id, current_user, db, owner_only=True)
    tournament.deleted_at = None
    db.commit()


@app.delete("/tournaments/{tournament_id}/permanent", status_code=204)
def permanently_delete_tournament(
    tournament_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """完全削除（ゴミ箱からのみ）"""
    tournament = db.query(Tournament).filter(
        Tournament.id == tournament_id, Tournament.deleted_at != None
    ).first()
    if tournament is None:
        raise HTTPException(status_code=404, detail="Tournament not found in trash")
    check_tournament_access(tournament_id, current_user, db, owner_only=True)

    for race in db.query(Race).filter(Race.tournament_id == tournament_id).all():
        db.query(RaceResult).filter(RaceResult.race_id == race.id).delete()
    db.query(Race).filter(Race.tournament_id == tournament_id).delete()
    db.query(Boat).filter(Boat.tournament_id == tournament_id).delete()
    db.query(TournamentMember).filter(TournamentMember.tournament_id == tournament_id).delete()
    db.query(RuleConfig).filter(RuleConfig.tournament_id == tournament_id).delete()
    db.delete(tournament)
    db.commit()


# ─── 認証・管理者エンドポイント ────────────────────────────────────────────

@app.post("/admin/bootstrap")
def bootstrap_admin(db: Session = Depends(get_db)):
    """
    初回セットアップ用：BOOTSTRAP_ADMIN_EMAIL のユーザーを admin に昇格する。
    - BOOTSTRAP_ADMIN_EMAIL 環境変数が未設定なら 400
    - 既に admin ユーザーが存在すれば 409（二重実行防止）
    - 対象メールのユーザーが users テーブルに未登録なら 404
    """
    email = os.environ.get("BOOTSTRAP_ADMIN_EMAIL", "").strip()
    if not email:
        raise HTTPException(status_code=400, detail="BOOTSTRAP_ADMIN_EMAIL 環境変数が設定されていません")

    existing_admin = db.query(User).filter(User.role == "admin").first()
    if existing_admin:
        raise HTTPException(status_code=409, detail="管理者ユーザーが既に存在します（二重実行防止）")

    user = db.query(User).filter(User.email == email).first()
    if user is None:
        raise HTTPException(
            status_code=404,
            detail=f"{email} のユーザーレコードが見つかりません。先に一度ログインしてください。",
        )

    user.role = "admin"
    db.commit()
    return {"message": f"{email} を admin に昇格しました", "user_id": user.id}


@app.get("/auth/me", response_model=UserOut)
def get_me(current_user=Depends(get_current_user)):
    """ログイン中のユーザー情報を返す"""
    return current_user


@app.get("/admin/users", response_model=list[UserOut])
def list_users(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    """全ユーザー一覧（管理者のみ）"""
    return db.query(User).order_by(User.email).all()


@app.post("/admin/invite", response_model=InviteResponse)
def invite_user(
    body: InviteRequest,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    """メールで招待を送り、usersテーブルに登録する（管理者のみ）"""
    if not AUTH_ENABLED:
        raise HTTPException(status_code=503, detail="認証が無効です（SUPABASE_URL 未設定）")

    try:
        supabase = get_supabase()
        resp = supabase.auth.admin.invite_user_by_email(body.email)
        sb_user = resp.user
        if sb_user is None:
            raise HTTPException(status_code=500, detail="Supabase 招待に失敗しました")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"招待エラー: {e}")

    # users テーブルに登録（既存なら role だけ更新）
    user = db.query(User).filter(User.id == sb_user.id).first()
    if user is None:
        user = User(id=sb_user.id, email=body.email, role=body.role)
        db.add(user)
    else:
        user.role = body.role

    # member の場合は担当大会を登録
    if body.role == "member":
        for tid in body.tournament_ids:
            exists = db.query(TournamentMember).filter(
                TournamentMember.tournament_id == tid,
                TournamentMember.user_id == sb_user.id,
            ).first()
            if not exists:
                db.add(TournamentMember(tournament_id=tid, user_id=sb_user.id))

    db.commit()
    return InviteResponse(message="招待メールを送信しました", email=body.email, role=body.role)


@app.put("/admin/users/{user_id}/role")
def update_user_role(
    user_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    """ユーザーの role を admin ↔ member 切り替え（管理者のみ）。自分自身は変更不可。"""
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="自分自身の権限は変更できません")

    target = db.query(User).filter(User.id == user_id).first()
    if target is None:
        raise HTTPException(status_code=404, detail="ユーザーが見つかりません")

    target.role = "member" if target.role == "admin" else "admin"
    db.commit()
    return {"id": target.id, "email": target.email, "role": target.role}


@app.post("/admin/users/{user_id}/tournaments/{tournament_id}")
def add_tournament_member(
    user_id: str,
    tournament_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    """ユーザーに担当大会を追加（管理者のみ）"""
    exists = db.query(TournamentMember).filter(
        TournamentMember.tournament_id == tournament_id,
        TournamentMember.user_id == user_id,
    ).first()
    if not exists:
        db.add(TournamentMember(tournament_id=tournament_id, user_id=user_id))
        db.commit()
    return {"message": "追加しました"}
